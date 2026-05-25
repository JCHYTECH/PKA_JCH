#!/usr/bin/env python3
# tests/procurement/fixtures/generate_fixture_xlsx.py
import openpyxl
from pathlib import Path

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOM"

# Ligne titre fusionnée (comme BOM WildNexus réelle)
ws.merge_cells("A1:G1")
ws["A1"] = "WildNexus P0 — Test BOM | Mis à jour : 2026-05-25"

headers = ["Ref", "Description", "Qty", "MPN", "Preferred Supplier", "Category", "Notes"]
ws.append(headers)

rows = [
    ("U1", "ESP32-S3-WROOM-1U-N8R8", 5, "ESP32-S3-WROOM-1U-N8R8", "Mouser", "MCU", "8MB PSRAM"),
    ("C1", "Condensateur 100nF 0402", 20, "", "Mouser", "Passif", ""),
    ("U2", "SHT31-DIS-B", 2, "SHT31-DIS-B2.5kS", "Mouser", "Capteur", "I2C"),
]
for row in rows:
    ws.append(row)

out = Path(__file__).parent / "sample_bom.xlsx"
wb.save(out)
print(f"Fixture créée : {out}")
