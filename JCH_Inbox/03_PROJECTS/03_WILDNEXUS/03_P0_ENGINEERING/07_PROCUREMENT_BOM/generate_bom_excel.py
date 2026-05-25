"""
Génère le fichier Excel WildNexus BOM P0 depuis les données de la shortlist v0.1
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(color="FFFFFF", bold=True, size=10)
SUB_FILL   = PatternFill("solid", fgColor="2E75B6")
SUB_FONT   = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL   = PatternFill("solid", fgColor="D6E4F0")
WARN_FILL  = PatternFill("solid", fgColor="FFF2CC")
OK_FILL    = PatternFill("solid", fgColor="E2EFDA")
RED_FILL   = PatternFill("solid", fgColor="FCE4D6")
TOTAL_FONT = Font(bold=True, size=10)
THIN       = Side(style="thin", color="BBBBBB")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP       = Alignment(wrap_text=True, vertical="top")
CENTER     = Alignment(horizontal="center", vertical="top")

def style_header(cell, sub=False):
    cell.fill   = SUB_FILL if sub else HDR_FILL
    cell.font   = SUB_FONT if sub else HDR_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

def style_cell(cell, alt=False, warn=False, ok=False, red=False, bold=False, center=False):
    if warn:   cell.fill = WARN_FILL
    elif ok:   cell.fill = OK_FILL
    elif red:  cell.fill = RED_FILL
    elif alt:  cell.fill = ALT_FILL
    cell.font      = Font(bold=bold, size=10)
    cell.alignment = CENTER if center else WRAP
    cell.border    = BORDER

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

# ════════════════════════════════════════════════════════════════════════════
# Feuille 1 — Option A (Banc rapide)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Option A — Banc rapide"
freeze(ws, "C3")

ws.merge_cells("A1:L1")
title = ws["A1"]
title.value = "WildNexus P0 — Option A : Banc rapide  |  Mis à jour : 2026-05-25"
title.font  = Font(bold=True, size=12, color="1F4E79")
title.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 22

headers = ["#", "Composant", "Référence précise", "Fournisseur", "Qté",
           "Prix unit. TTC (est.)", "Total (est.)", "Délai", "Risque supply",
           "Statut PIR", "Note", "Action requise"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=c, value=h)
    style_header(cell)
ws.row_dimensions[2].height = 30

option_a = [
    ("A01", "MCU prototype",           "ESP32-S3-DevKitC-1-N8R8 (8MB flash, 8MB PSRAM, USB-C)",
     "Mouser BE",          2,  15.17,  30.34,  "3–5 j",  "🟢 Bas",   "✅ Confirmé",
     "Acheter 2 pour banc M-02 redondant",              ""),
    ("A02", "Caméra capteur",           "OV5640 5MP DVP M12 — Arducam B0253",
     "Arducam officiel",   2,  22.00,  44.00,  "5–10 j", "🟡 Moyen", "✅ Confirmé",
     "Vérifier DVP explicite + NoIR + M12 avant achat", "Contacter Arducam avant paiement"),
    ("A03", "LED IR 850 nm",            "Vishay VSLY3850 (×4/nœud)",
     "Mouser BE",         10,   0.80,   8.00,  "3–5 j",  "🟢 Bas",   "✅ Confirmé",
     "Test terrain 850 nm en premier",                  ""),
    ("A04", "LED IR 940 nm",            "Würth WL-SIRW 940 nm / Vishay VSLB3940",
     "Mouser BE",         10,   0.90,   9.00,  "3–5 j",  "🟢 Bas",   "✅ Confirmé",
     "Test comparatif 850/940 nm — ne pas figer avant mesures", ""),
    ("A05", "PIR basse conso",          "Panasonic EKMB1303111K (1×) + AM312 module (1×)",
     "Mouser BE / marketplace", 2, 11.00, 22.00, "5–8 j", "🟢 Bas",  "✅ Validé JCH 2026-05-25",
     "EKMB (~20€) pour mesures énergie + AM312 (~2€) pour tests flux", ""),
    ("A06", "Micro MEMS I2S",           "Knowles SPH0645LM4H-B (OBSOLÈTE — P0 OK, pas PCB final)",
     "DigiKey BE",         2,   3.15,   6.30,  "3–7 j",  "🟢 Bas",   "⚠️ Obsolète",
     "Utiliser P0 uniquement. Remplacer ICS-43434 ou MP34DT05TR-A pour PCB final", ""),
    ("A07", "microSD industrielle",     "Swissbit S-45U 16 GB (pSLC, −40/+85°C)",
     "DigiKey BE",         3,  24.00,  72.00,  "3–7 j",  "🟡 Moyen", "✅ Confirmé",
     "NE PAS acheter consumer. Alt: Transcend TS16GUSD350V",           ""),
    ("A08", "Buck faible Iq",           "TPS62840DLCR (Iq 60 nA, 2,5–5,5V → 3V, 750mA)",
     "Mouser BE",          5,   2.00,  10.00,  "3–5 j",  "🟢 Bas",   "✅ Confirmé",
     "DigiKey épuisé → Mouser uniquement (27 639 unités)",             ""),
    ("A09", "Holder AA × 4",            "Keystone 2463 (4× AA, 6V, fils) — 2 holders/nœud",
     "RS Belgium / Mouser BE", 4, 3.50, 14.00, "3–7 j",  "🟢 Bas",   "✅ Confirmé",
     "2 holders × 4 AA = 8 AA en 2 blocs parallèles (4S2P)",          ""),
    ("A10", "Inductance buck",          "Würth 744043100 (10 µH, 1 A, 3×3 mm)",
     "Mouser BE",          5,   0.60,   3.00,  "3–5 j",  "🟢 Bas",   "✅ Confirmé",
     "Associée TPS62840",                                              ""),
    ("A11", "Passifs CMS",              "Lot 0402/0603 : C 100nF/10µF/1µF ; R 10K/100K/1M",
     "Mouser BE",          1,  20.00,  20.00,  "3–5 j",  "🟢 Bas",   "✅ Confirmé",
     "Lot prototypage — couvre buck, micro, PIR, LED driver",          ""),
    ("A12", "MOSFET power gating",      "Nexperia PMV16XN (N-ch, SOT-23) ou SI2302",
     "Mouser BE",          5,   0.50,   2.50,  "3–5 j",  "🟢 Bas",   "✅ Confirmé",
     "Power gating caméra/IR/LoRa",                                    ""),
    ("A13", "Nordic PPK2",              "Nordic Power Profiler Kit II — NON NÉGOCIABLE",
     "Nordic Semi EU / Mouser BE", 1, 110.34, 110.34, "5–10 j", "🟢 Bas", "✅ Confirmé",
     "Sans PPK2 → ADR-005 non vérifiable. Priorité absolue.",          "Commander en premier"),
    ("A14", "Breadboard + câbles",      "Kit proto 830 pts + câbles Dupont",
     "Amazon.de / Conrad", 2,   7.00,  14.00,  "3–5 j",  "🟢 Bas",   "✅ Confirmé",
     "Banc ouvert M-02",                                               ""),
    ("A15", "Alimentation labo",        "RD UM24C ou Ruideng RD6006",
     "Amazon.de",          1,  35.00,  35.00,  "3–7 j",  "🟢 Bas",   "✅ Confirmé",
     "Si alimentation labo inexistante chez JCH",                      ""),
    ("A16", "Piles AA lithium",         "Energizer Ultimate Lithium L91 AA × 16",
     "Carrefour / Amazon.be", 16, 1.80, 28.80, "Immédiat", "🟢 Bas",  "✅ Confirmé",
     "16 piles = 2 nœuds × 8 AA. Lithium = test autonomie crédible froid", ""),
]

total_a = 0
for r, row in enumerate(option_a, 3):
    alt = (r % 2 == 0)
    warn = "⚠️" in str(row[10]) or "OBSOLÈTE" in str(row[2])
    ok   = "✅" in str(row[9])
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        style_cell(cell, alt=alt, warn=warn, ok=(ok and not warn))
    total_a += row[6]
    ws.row_dimensions[r].height = 40

# Sous-total
r_sub = len(option_a) + 3
ws.merge_cells(f"A{r_sub}:F{r_sub}")
sub = ws[f"A{r_sub}"]
sub.value = "Sous-total Option A (prix réels vérifiés)"
sub.font  = TOTAL_FONT
sub.fill  = PatternFill("solid", fgColor="2E75B6")
sub.font  = Font(bold=True, color="FFFFFF")
sub.alignment = Alignment(horizontal="right", vertical="center")
tot = ws.cell(row=r_sub, column=7, value=f"~{total_a:.0f} €")
tot.font = Font(bold=True, color="FFFFFF")
tot.fill = PatternFill("solid", fgColor="2E75B6")
tot.alignment = CENTER

set_col_widths(ws, [5, 20, 35, 18, 5, 14, 12, 8, 10, 18, 45, 30])

# ════════════════════════════════════════════════════════════════════════════
# Feuille 2 — Option B (Terrain compact)
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Option B — Terrain compact")
freeze(ws2, "C3")

ws2.merge_cells("A1:L1")
title2 = ws2["A1"]
title2.value = "WildNexus P0 — Option B : Additions terrain compact  |  À commander après résultats M-02 banc"
title2.font  = Font(bold=True, size=12, color="1F4E79")
title2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 22

for c, h in enumerate(headers, 1):
    cell = ws2.cell(row=2, column=c, value=h)
    style_header(cell)
ws2.row_dimensions[2].height = 30

option_b = [
    ("B01", "Lentille M12 IR-corrigée", "Lensation LS-6028 f/1.8 6mm IR / Evetar M12B0618W-IR",
     "Lensation.de",       2,  40.00,  80.00,  "5–10 j", "🟡 Moyen", "⚠️ Ne pas < 30€/lentille",
     "Qualité nuit critique. Demander devis 2 pièces avant toute commande.", "Devis Lensation obligatoire"),
    ("B02", "Module LoRa nœud",         "RAK3172-T EU868 (SX1262, CE certifié, IPEX, −40/+85°C)",
     "RAK Store EU",        2,   8.00,  16.00,  "5–10 j", "🟢 Bas",   "✅ Confirmé",
     "Vérifier variante EU868 + IPEX. Alt: Murata Type 1SJ si rupture RAK.", ""),
    ("B03", "Antenne LoRa 868 MHz",     "LoRa flex antenna 868 MHz IPEX — Molex 0213980100 ~80 mm",
     "Mouser BE",           2,   4.00,   8.00,  "3–7 j",  "🟢 Bas",   "✅ Confirmé",
     "Antenne interne. Alt: câble pigtail IPEX→SMA + antenne camouflée branche.", ""),
    ("B04", "Boîtier IP67 P0",          "Hammond 1554A (150×100×60 mm) ou 1554B (188×120×78 mm) ABS IP67",
     "RS Belgium / Mouser BE", 2, 16.00, 32.00, "3–7 j",  "🟡 Moyen", "⚠️ Après layout A",
     "Commander après layout interne Option A validé. P1 = impression 3D forme organique dos arrondi (décision JCH 2026-05-25). Hors budget P0.", "Attendre layout A"),
    ("B05", "BME688 capteurs env.",      "Bosch BME688 breakout ou puce CMS",
     "Mouser BE",           2,   9.00,  18.00,  "3–7 j",  "🟢 Bas",   "✅ Optionnel P0",
     "Réserve mécanique et firmware. Ne pas complexifier M-01/M-02 si budget serré.", ""),
    ("B06", "SHT31-DIS humidité int.",  "Sensirion SHT31-DIS-B (±2% RH, ±0.3°C, I2C)",
     "Mouser BE",           2,   5.00,  10.00,  "3–5 j",  "🟢 Bas",   "✅ Confirmé",
     "Détection condensation interne / silica gel saturé.",            ""),
    ("B07", "Connecteur étanche",       "Conxall Mini-Con-X 3 brins ou presse-étoupe M16 IP68",
     "RS Belgium / Mouser", 4,   2.00,   8.00,  "3–7 j",  "🟢 Bas",   "✅ Confirmé",
     "Passage câble sangle/antenne externe si antenne fouet obligatoire.", ""),
    ("B08", "Silica gel + joints",      "Sachets 5g silica gel ×10 + joint torique EPDM 140×3",
     "Conrad / Amazon.be",  1,  10.00,  10.00,  "Immédiat","🟢 Bas",   "✅ Confirmé",
     "Maintenance terrain, étanchéité boîtier IP67.",                  ""),
    ("B09", "Sangle UV terrain",        "Sangle polyester UV 20mm × 50cm + boucle inox",
     "Conrad / Camping",    4,   4.00,  16.00,  "Immédiat","🟢 Bas",   "✅ Confirmé",
     "Fixation sur tronc.",                                            ""),
]

total_b = 0
for r, row in enumerate(option_b, 3):
    alt  = (r % 2 == 0)
    warn = "⚠️" in str(row[9]) or "⚠️" in str(row[10])
    ok   = "✅" in str(row[9])
    for c, val in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        style_cell(cell, alt=alt, warn=warn, ok=(ok and not warn))
    total_b += row[6]
    ws2.row_dimensions[r].height = 45

r_sub2 = len(option_b) + 3
ws2.merge_cells(f"A{r_sub2}:F{r_sub2}")
sub2 = ws2[f"A{r_sub2}"]
sub2.value = "Sous-total Option B"
sub2.font  = Font(bold=True, color="FFFFFF")
sub2.fill  = PatternFill("solid", fgColor="2E75B6")
sub2.alignment = Alignment(horizontal="right", vertical="center")
tot2 = ws2.cell(row=r_sub2, column=7, value=f"~{total_b:.0f} €")
tot2.font = Font(bold=True, color="FFFFFF")
tot2.fill = PatternFill("solid", fgColor="2E75B6")
tot2.alignment = CENTER

set_col_widths(ws2, [5, 20, 35, 18, 5, 14, 12, 8, 10, 20, 45, 25])

# ════════════════════════════════════════════════════════════════════════════
# Feuille 3 — Budget consolidé
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Budget consolidé")

ws3.merge_cells("A1:D1")
t3 = ws3["A1"]
t3.value = "WildNexus P0 — Budget consolidé  |  Enveloppe : 1 000 €  |  2026-05-25"
t3.font  = Font(bold=True, size=12, color="1F4E79")
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 22

budget_rows = [
    ("Bloc", "Estimé initial", "Révisé (prix réels)", "Note", True),
    ("Option A — Banc rapide", "396 €", "~425 €", "Prix vérifiés Mouser/DigiKey/Nordic", False),
    ("Option B — Additions terrain", "203 €", "~203 €", "Prix estimés — à vérifier avant commande B", False),
    ("Frais de port (4–5 colis)", "50 €", "~50 €", "Mouser + DigiKey + RAK + Lensation + Nordic", False),
    ("Marge casse / reprise", "80 €", "~80 €", "", False),
    ("TOTAL estimé A + B", "729 €", "~758 €", "", True),
    ("Réserve restante (enveloppe 1 000 €)", "271 €", "~242 €", "✅ Budget tenu", True),
]

for r, (label, est, rev, note, bold) in enumerate(budget_rows, 2):
    is_hdr = bold and r == 2
    for c, val in enumerate([label, est, rev, note], 1):
        cell = ws3.cell(row=r, column=c, value=val)
        if is_hdr:
            style_header(cell)
        elif bold and "TOTAL" in label:
            cell.font = TOTAL_FONT
            cell.fill = PatternFill("solid", fgColor="2E75B6")
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = CENTER
            cell.border = BORDER
        elif bold and "Réserve" in label:
            cell.font = Font(bold=True, size=10)
            cell.fill = OK_FILL
            cell.alignment = WRAP
            cell.border = BORDER
        else:
            alt = (r % 2 == 0)
            style_cell(cell, alt=alt)
    ws3.row_dimensions[r].height = 22

# Règle dépassement
ws3.row_dimensions[len(budget_rows)+3].height = 18
warn_cell = ws3.cell(row=len(budget_rows)+3, column=1,
    value="⚠️  Tout achat individuel > 50 € → signaler à JCH avant commande.   |   Total projeté > 900 € → stopper et arbitrer.")
warn_cell.font = Font(bold=True, color="C00000", size=10)
ws3.merge_cells(f"A{len(budget_rows)+3}:D{len(budget_rows)+3}")

set_col_widths(ws3, [35, 16, 20, 45])

# ════════════════════════════════════════════════════════════════════════════
# Feuille 4 — ADR Status & Paniers
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("ADR & Paniers")

ws4.merge_cells("A1:F1")
t4 = ws4["A1"]
t4.value = "WildNexus P0 — Statut ADR & Stratégie paniers fournisseurs"
t4.font  = Font(bold=True, size=12, color="1F4E79")
t4.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 22

adr_headers = ["ADR", "Sujet", "Statut", "Impact shortlist", "Lignes BOM"]
for c, h in enumerate(adr_headers, 1):
    cell = ws4.cell(row=2, column=c, value=h)
    style_header(cell)

adr_rows = [
    ("ADR-001", "ESP32-S3",            "✅ Accepté", "A01 — DevKit", "A01"),
    ("ADR-002", "Caméra OV5640/IR",    "✅ Accepté", "A02 module + B01 lentille", "A02, B01"),
    ("ADR-003", "LoRa P2P EU868",      "✅ Accepté", "B02 RAK3172-T + B03 antenne", "B02, B03"),
    ("ADR-004", "microSD industrielle","✅ Accepté", "A07 Swissbit DigiKey", "A07"),
    ("ADR-005", "AA + buck faible Iq", "✅ Accepté", "A08–A10 buck/inductance + A09 holders + A16 piles", "A08–A10, A16"),
    ("ADR-006", "Boîtier IP67",        "✅ Accepté (2026-05-25)", "B04 Hammond P0 ; impression 3D organique P1 hors budget", "B04"),
    ("ADR-007", "Détection PIR",       "✅ Accepté", "A05 EKMB + AM312 (validé JCH 2026-05-25)", "A05"),
    ("ADR-008", "Capteurs extensible", "✅ Accepté", "A06 micro + B05–B06 capteurs env.", "A06, B05, B06"),
    ("ADR-009", "Architecture Sat/Base/Cloud", "✅ Accepté", "Pas d'achat Pi/cloud P0", "—"),
]

for r, row in enumerate(adr_rows, 3):
    alt = (r % 2 == 0)
    for c, val in enumerate(row, 1):
        cell = ws4.cell(row=r, column=c, value=val)
        style_cell(cell, alt=alt, ok=True)
    ws4.row_dimensions[r].height = 30

# Séparateur
ws4.row_dimensions[len(adr_rows)+4].height = 8

# Paniers fournisseurs
prow = len(adr_rows) + 5
ws4.merge_cells(f"A{prow}:F{prow}")
ph = ws4[f"A{prow}"]
ph.value = "STRATÉGIE PANIERS FOURNISSEURS"
ph.font  = Font(bold=True, size=11, color="FFFFFF")
ph.fill  = PatternFill("solid", fgColor="1F4E79")
ph.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[prow].height = 20

panier_headers = ["Panier", "Fournisseur", "Contenu", "Seuil livraison gratuite", "Priorité", "Action"]
for c, h in enumerate(panier_headers, 1):
    cell = ws4.cell(row=prow+1, column=c, value=h)
    style_header(cell, sub=True)

paniers = [
    ("P1", "Mouser BE",       "A01 MCU, A03–A12 capteurs/passifs/buck/MOSFET, B03 antenne, B05–B06 capteurs env.", "~50 € (à confirmer)", "🔴 En premier", ""),
    ("P2", "DigiKey BE",      "A07 microSD Swissbit, A13 Nordic PPK2", "≥50 € livraison gratuite", "🔴 En premier", ""),
    ("P3", "Arducam officiel","A02 module caméra OV5640 DVP M12", "Livraison directe EU", "🔴 En premier", "Contacter Arducam avant paiement"),
    ("P4", "Nordic Semi EU",  "A13 PPK2 (si pas Mouser)", "Livraison directe EU", "🔴 En premier", "Commander en même temps que P1"),
    ("P5", "RAK Store EU",    "B02 RAK3172-T EU868", "Livraison directe EU", "🟡 Après M-02 banc", ""),
    ("P6", "Lensation.de",    "B01 lentille M12 f/1.8 IR-corrigée × 2", "Demander devis", "🟡 Après M-02 banc", "Devis 2 pièces obligatoire"),
    ("P7", "RS Belgium",      "A09 holders AA, B04 boîtier Hammond, B07 presse-étoupe", "Livraison rapide BE", "🟡 Après layout A validé", ""),
]

for r, row in enumerate(paniers, prow+2):
    alt = (r % 2 == 0)
    warn = "Devis" in str(row[5]) or "avant paiement" in str(row[5])
    for c, val in enumerate(row, 1):
        cell = ws4.cell(row=r, column=c, value=val)
        style_cell(cell, alt=alt, warn=warn)
    ws4.row_dimensions[r].height = 35

set_col_widths(ws4, [10, 18, 50, 22, 18, 30])

# ════════════════════════════════════════════════════════════════════════════
# Sauvegarde
# ════════════════════════════════════════════════════════════════════════════
OUT = "/Users/jchavauxm5/PKA_JCH/JCH_Inbox/03_PROJECTS/03_WILDNEXUS/03_P0_ENGINEERING/07_PROCUREMENT_BOM/WILDNEXUS_BOM_P0_v0.2.xlsx"
wb.save(OUT)
print(f"✅ Fichier généré : {OUT}")
