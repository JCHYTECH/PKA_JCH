# scripts/procurement/bom_writer.py
import csv
import shutil
from datetime import date
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from .models import ProcurementResult, EnrichedComponent

_ENRICH_HEADERS = [
    "Prix unit. (EUR)", "Stock", "Délai (j)", "Obsolète",
    "Alternative", "URL fournisseur", "Simulation",
]

_BLUE_DARK = "1F4E79"
_BLUE_LIGHT = "D6E4F0"
_YELLOW = "FFF2CC"
_RED = "FFD7D7"


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _enrich_row(e: EnrichedComponent) -> list:
    return [
        f"{e.price_unit:.2f}" if e.price_unit is not None else "N/A",
        str(e.stock_qty) if e.stock_qty is not None else "N/A",
        str(e.lead_time_days) if e.lead_time_days is not None else "N/A",
        "🔴 OUI" if e.obsolete else "NON",
        ", ".join(e.alternatives) if e.alternatives else "",
        e.supplier_url,
        "OUI" if e.simulation else "NON",
    ]


def _enrich_xlsx(src: Path, result: ProcurementResult, output_dir: Path) -> Path:
    out_name = src.stem + f"_enrichi_{date.today().isoformat()}.xlsx"
    out = output_dir / out_name
    shutil.copy(src, out)

    wb = openpyxl.load_workbook(out)
    ws = next((wb[n] for n in wb.sheetnames if "bom" in n.lower()), wb.active)

    # Trouver la ligne de headers (première ligne avec > 1 valeur non-None)
    header_row_idx = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        non_null = [v for v in row if v is not None and str(v).strip()]
        if len(non_null) > 1:
            header_row_idx = idx
            break

    if header_row_idx is None:
        wb.save(out)
        return out

    last_col = ws.max_column
    for i, h in enumerate(_ENRICH_HEADERS):
        cell = ws.cell(row=header_row_idx, column=last_col + 1 + i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _fill(_BLUE_DARK)
        cell.alignment = Alignment(wrap_text=True)

    enriched_map = {e.component.ref: e for e in result.enriched}

    headers_raw = [ws.cell(row=header_row_idx, column=c).value
                   for c in range(1, last_col + 1)]
    ref_col = next(
        (i + 1 for i, h in enumerate(headers_raw)
         if h and "ref" in str(h).lower()),
        1
    )
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        ref_val = ws.cell(row=row_idx, column=ref_col).value
        if not ref_val:
            continue
        e = enriched_map.get(str(ref_val).strip())
        if not e:
            continue
        values = _enrich_row(e)
        row_fill = _fill(_RED if e.obsolete else (_YELLOW if e.simulation else _BLUE_LIGHT))
        for i, val in enumerate(values):
            cell = ws.cell(row=row_idx, column=last_col + 1 + i, value=val)
            cell.fill = row_fill

    wb.save(out)
    return out


def _enrich_csv(src: Path, result: ProcurementResult, output_dir: Path) -> Path:
    out_name = src.stem + f"_enrichi_{date.today().isoformat()}.csv"
    out = output_dir / out_name
    enriched_map = {e.component.ref: e for e in result.enriched}

    with open(src, encoding="utf-8") as fin, open(out, "w", newline="", encoding="utf-8") as fout:
        reader = csv.DictReader(fin)
        fieldnames = list(reader.fieldnames or []) + _ENRICH_HEADERS
        writer = csv.DictWriter(fout, fieldnames=fieldnames)
        writer.writeheader()
        for row in reader:
            ref = row.get("ref", row.get("Ref", "")).strip()
            e = enriched_map.get(ref)
            if e:
                values = _enrich_row(e)
                row.update(dict(zip(_ENRICH_HEADERS, values)))
            writer.writerow(row)

    return out


def enrich_bom(result: ProcurementResult, output_dir: Path) -> Path:
    src = Path(result.bom.source_file)
    output_dir.mkdir(parents=True, exist_ok=True)
    if src.suffix.lower() in (".xlsx", ".xls"):
        return _enrich_xlsx(src, result, output_dir)
    else:
        return _enrich_csv(src, result, output_dir)
