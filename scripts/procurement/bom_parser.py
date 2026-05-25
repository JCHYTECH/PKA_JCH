# scripts/procurement/bom_parser.py
import csv
import re
from pathlib import Path
import openpyxl
from .models import BOMData, Component

# Mapping noms de colonnes → champs Component
_COLUMN_MAP = {
    "ref": ["ref", "reference", "designator", "repère"],
    "description": ["description", "desc", "composant", "component", "désignation"],
    "qty": ["qty", "quantity", "quantité", "qte", "q", "nb"],
    "mpn": ["mpn", "part number", "manufacturer part number", "référence fabricant", "pn"],
    "preferred_supplier": ["preferred_supplier", "fournisseur", "supplier", "preferred supplier"],
    "category": ["category", "catégorie", "type", "famille"],
    "notes": ["notes", "note", "remarques", "commentaire"],
}


def _map_column(header: str) -> str:
    h = header.strip().lower()
    for field, aliases in _COLUMN_MAP.items():
        if h in aliases:
            return field
    return h


def _row_to_component(row: dict) -> Component:
    try:
        qty = int(float(str(row.get("qty", 1) or 1)))
    except (ValueError, TypeError):
        qty = 1
    return Component(
        ref=str(row.get("ref", "")).strip(),
        description=str(row.get("description", "")).strip(),
        qty=qty,
        mpn=str(row.get("mpn", "") or "").strip(),
        preferred_supplier=str(row.get("preferred_supplier", "") or "").strip(),
        category=str(row.get("category", "") or "").strip(),
        notes=str(row.get("notes", "") or "").strip(),
    )


def _is_title_row(row: tuple) -> bool:
    """Ignore les lignes de titre fusionnées (1 seule valeur non-None)."""
    non_null = [v for v in row if v is not None and str(v).strip()]
    return len(non_null) <= 1


def parse_excel(path: Path, project: str) -> BOMData:
    wb = openpyxl.load_workbook(path)
    # Cherche une feuille nommée BOM, sinon prend la première
    ws = next((wb[n] for n in wb.sheetnames if "bom" in n.lower()), wb.active)

    headers = None
    components = []
    for row in ws.iter_rows(values_only=True):
        if all(v is None for v in row):
            continue
        if headers is None:
            if _is_title_row(row):
                continue
            headers = [_map_column(str(v)) if v is not None else "" for v in row]
            continue
        row_dict = {h: v for h, v in zip(headers, row)}
        if not row_dict.get("ref") and not row_dict.get("description"):
            continue
        components.append(_row_to_component(row_dict))

    return BOMData(project=project, source_file=str(path), components=components)


def parse_csv(path: Path, project: str) -> BOMData:
    with open(path, encoding="utf-8") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;")
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(f, dialect=dialect)
        mapped_headers = {k: _map_column(k) for k in (reader.fieldnames or [])}
        components = []
        for row in reader:
            mapped = {mapped_headers[k]: v for k, v in row.items()}
            if not mapped.get("ref") and not mapped.get("description"):
                continue
            components.append(_row_to_component(mapped))

    return BOMData(project=project, source_file=str(path), components=components)


def parse_markdown(path: Path, project: str) -> BOMData:
    content = path.read_text(encoding="utf-8")
    headers = None
    components = []
    for line in content.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        # Ligne séparateur (---)
        if all(re.match(r"^:?-+:?$", c) or c == "" for c in cells):
            continue
        if headers is None:
            headers = [_map_column(c) for c in cells]
            continue
        row_dict = dict(zip(headers, cells))
        if not row_dict.get("ref") and not row_dict.get("description"):
            continue
        components.append(_row_to_component(row_dict))

    return BOMData(project=project, source_file=str(path), components=components)


def parse(path: Path, project: str) -> BOMData:
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xls"):
        return parse_excel(path, project)
    elif ext == ".csv":
        return parse_csv(path, project)
    elif ext == ".md":
        return parse_markdown(path, project)
    else:
        raise ValueError(
            f"Format non supporté: {ext}. Formats acceptés: .xlsx, .csv, .md"
        )
